"""
Command-line interface for DirectorSync v3 (Step 10, minimal).

Usage (examples):
  - Dry-run (no HTTP):
      python -m directorsync_v3.cli apply --profile repos --rows ./data/example.csv --dry-run

  - Real apply (HTTP CRUD):
      python -m directorsync_v3.cli apply --profile repos --rows ./data/example.csv \
        --base-url http://127.0.0.1:8000 --token TEST
"""

from __future__ import annotations

import argparse
import os
import csv
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List

from .core.config import load_config
from .core.logging_setup import build_logger
from .core.profiles import ProfileLoader
from .core.importer import GenericImporter
from .core.director_client import DirectorClient
from .core.applier import CrudApplier


# --- add this helper next to _read_rows_csv ---
def _read_rows_xlsx(path: str) -> List[Dict[str, Any]]:
    """
    Minimal XLSX reader using openpyxl (optional dependency).
    - Uses the first worksheet
    - First row = headers
    - Returns a list of {header: string_value}
    """
    try:
        from openpyxl import load_workbook  # lazy import so dependency stays optional
    except Exception as e:
        raise RuntimeError(
            "XLSX support requires 'openpyxl'. Install it (e.g. `pip install openpyxl`) "
            "or use a CSV file instead."
        ) from e

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Rows file not found: {path}")

    wb = load_workbook(filename=str(p), data_only=True, read_only=True)
    ws = wb.active  # first sheet
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [("" if h is None else str(h).strip()) for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        d: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = r[i] if i < len(r) else None
            d[h] = "" if val is None else str(val)
        out.append(d)
    return out

def _read_rows_auto(path: str) -> List[Dict[str, Any]]:
    """
    Auto-detect reader by file extension.
    """
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_rows_xlsx(path)
    # default to CSV
    return _read_rows_csv(path)

def _read_rows_csv(path: str) -> List[Dict[str, Any]]:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Rows file not found: {path}")
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]


def _summarize_counts(counts: Dict[str, int]) -> str:
    # stable order for readability
    keys = ["CREATED", "UPDATED", "UNCHANGED", "SKIP", "ERROR", "EXCEPTION"]
    parts = [f"{k}={counts.get(k, 0)}" for k in keys]
    return " | ".join(parts)


def _exit_code_from_counts(counts: Dict[str, int]) -> int:
    if counts.get("ERROR", 0) or counts.get("EXCEPTION", 0):
        return 2
    return 0


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="dsync", description="DirectorSync v3 CLI (minimal)")

    sub = p.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("apply", help="Apply rows against a Director profile")
    a.add_argument("--profile", required=True, help="Profile name (without .yml)")
    a.add_argument("--rows", required=True, help="Input rows file (.csv or .xlsx)")
    a.add_argument("--search-path", default="resources/profiles", help="Profiles search path")

    a.add_argument("--dry-run", action="store_true", help="Plan only, no network calls")

    # Director / HTTP
    a.add_argument("--base-url", default="", help="Director base URL")
    a.add_argument("--token", default="", help="Director API token")
    a.add_argument("--verify-tls", default="true", choices=["true", "false"], help="Verify TLS (https)")
    a.add_argument("--timeout-sec", type=int, default=30, help="HTTP timeout seconds")
    a.add_argument("--retries", type=int, default=3, help="HTTP retries (5xx/network)")

    # Context
    a.add_argument("--tenant", default="", help="Context tenant")
    a.add_argument("--pool-uuid", default="", help="Context pool UUID")

    # Logging
    a.add_argument("--logs-dir", default="logs", help="Logs base directory")
    a.add_argument("--console-level", default="INFO", help="Console log level (INFO..CRITICAL)")
    a.add_argument("--file-level", default="DEBUG", help="File log level (DEBUG..CRITICAL)")

    return p


def _apply_cmd(args: argparse.Namespace) -> int:
    # 1) Build config from CLI overrides
    # Provide sane defaults for non-dry-run context if not passed on CLI
    # Defaults for required non-dry-run context
    _tenant = args.tenant or os.environ.get("DIRECTORSYNC_TENANT", "cli-tenant")
    _pool   = args.pool_uuid or os.environ.get("DIRECTORSYNC_POOL", "cli-pool")

    cli_overrides = {    

        "app": {"dry_run": bool(args.dry_run)},
        "director": {
            "base_url": args.base_url,
            "token": args.token,
            "verify_tls": (args.verify_tls.lower() == "true"),
            "timeout_sec": int(args.timeout_sec),
            "retries": int(args.retries),
        },
        "context": {"tenant": _tenant, "pool_uuid": _pool},
        "logging": {
            "base_dir": args.logs_dir,
            "console_level": args.console_level,
            "file_level": args.file_level,
        },
    }
    cfg = load_config(cli_overrides)

    # 2) Logger
    logger = build_logger(
        run_id=cfg.run_id,
        action="apply",
        base_dir=cfg.logging.base_dir,
        console_level=cfg.logging.console_level,
        file_level=cfg.logging.file_level,
        extra={"tenant": cfg.context.tenant, "pool": cfg.context.pool_uuid, "profile": args.profile},
    )
    logger.info("Starting dsync apply (dry_run=%s)", cfg.app.dry_run)

    # 3) Profile
    pl = ProfileLoader(search_paths=[args.search_path])
    profile = pl.load(args.profile)

    # 4) Read rows
    rows = _read_rows_auto(args.rows)
    
    logger.info("Loaded %s input rows from %s", len(rows), args.rows)

    # 5) Execute
    if cfg.app.dry_run:
        importer = GenericImporter(profile, logger=logger)
        results, counts = importer.run(rows, remote_items=[])
        logger.info("Dry-run summary: %s", _summarize_counts(counts))
        print(_summarize_counts(counts))
        return _exit_code_from_counts(counts)

    # Real apply (HTTP)
    client = DirectorClient(
        base_url=cfg.director.base_url,
        token=cfg.director.token,
        verify_tls=bool(cfg.director.verify_tls),
        timeout_sec=int(cfg.director.timeout_sec),
        retries=int(cfg.director.retries),
        logger=logger,
    )
    applier = CrudApplier(
        profile,
        client,
        context={"tenant": cfg.context.tenant, "pool_uuid": cfg.context.pool_uuid},
        logger=logger,
    )
    results, counts = applier.apply(rows)
    logger.info("Apply summary: %s", _summarize_counts(counts))
    print(_summarize_counts(counts))
    return _exit_code_from_counts(counts)


def main(argv: Iterable[str] | None = None) -> int:
    parser = _build_arg_parser()
    args = parser.parse_args(list(argv) if argv is not None else None)

    if args.cmd == "apply":
        return _apply_cmd(args)

    parser.error("Unknown command")  # pragma: no cover
    return 2  # pragma: no cover


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
